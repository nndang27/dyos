"""
Excel Report Generator
Creates a formatted Excel report from CSV with multiple sheets, 
conditional formatting, pivot table, and charts.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import os

# Read the CSV file
print("Reading CSV file...")
csv_path = '/Users/nndang27/Documents/nano_claw/plan_react_system/deepagents/test_folder/32130_2026A_10(in).csv'
df = pd.read_csv(csv_path)
print(f"Loaded {len(df)} records with {len(df.columns)} columns")
print(f"Columns: {list(df.columns)}")

# Ensure output directory exists
output_dir = '/Users/nndang27/Documents/nano_claw/plan_react_system/deepagents/test_folder/output'
os.makedirs(output_dir, exist_ok=True)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Demographics"

# Define column mappings for each sheet
sheet_configs = {
    "Demographics": ['source ID', 'age', 'gender', 'marital_status', 'education_level'],
    "Income": ['annual_income', 'monthly_income', 'employment_status'],
    "Credit": ['credit_score', 'debt_to_income_ratio', 'num_of_open_accounts', 
               'total_credit_limit', 'current_balance', 'delinquency_history', 'num_of_delinquencies'],
    "Loan": ['loan_amount', 'loan_purpose', 'interest_rate', 'loan_term', 'installment', 'loan_paid_back']
}

# Define numeric columns for outlier detection
numeric_columns = {
    "Demographics": ['age'],
    "Income": ['annual_income', 'monthly_income'],
    "Credit": ['credit_score', 'debt_to_income_ratio', 'num_of_open_accounts', 
               'total_credit_limit', 'current_balance', 'delinquency_history', 'num_of_delinquencies'],
    "Loan": ['loan_amount', 'interest_rate', 'loan_term', 'installment']
}

# Style definitions
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
outlier_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
outlier_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def calculate_iqr_outliers(series):
    """Calculate IQR-based outlier bounds"""
    Q1 = series.quantile(0.25)
    Q3 = series.quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    return lower_bound, upper_bound

def write_sheet_data(ws, df, columns, sheet_name):
    """Write data to sheet with formatting"""
    # Write header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Write data
    for row_idx, row in df[columns].iterrows():
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=row[col_name])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        for row in range(2, len(df) + 2):
            try:
                cell_value = str(ws.cell(row=row, column=col_idx).value)
                max_length = max(max_length, len(cell_value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    return ws

def apply_conditional_formatting(ws, df, columns, numeric_cols):
    """Apply conditional formatting for outliers"""
    for col_name in numeric_cols:
        if col_name in columns:
            col_idx = columns.index(col_name) + 1
            series = pd.to_numeric(df[col_name], errors='coerce').dropna()
            
            if len(series) > 0:
                lower_bound, upper_bound = calculate_iqr_outliers(series)
                
                # Apply conditional formatting for outliers
                # For values below lower bound
                for row_idx in range(2, len(df) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        value = float(cell.value)
                        if value < lower_bound or value > upper_bound:
                            cell.fill = outlier_fill
                            cell.font = outlier_font
                    except (ValueError, TypeError):
                        pass
    
    return ws

# Create each sheet
print("\nCreating sheets...")
for sheet_name, columns in sheet_configs.items():
    if sheet_name == "Demographics":
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    print(f"  - {sheet_name}")
    write_sheet_data(ws, df, columns, sheet_name)
    apply_conditional_formatting(ws, df, columns, numeric_columns.get(sheet_name, []))

# Create Summary Pivot Table sheet
print("\nCreating Summary Pivot Table...")
ws_pivot = wb.create_sheet(title="Summary")

# Summary statistics by loan_purpose
pivot_loan_purpose = df.groupby('loan_purpose').agg({
    'loan_amount': ['count', 'mean', 'min', 'max', 'sum'],
    'interest_rate': 'mean',
    'credit_score': 'mean'
}).round(2)

# Flatten column names
pivot_loan_purpose.columns = ['_'.join(col).strip() for col in pivot_loan_purpose.columns.values]

# Write pivot table header
pivot_columns = list(pivot_loan_purpose.columns)
ws_pivot.cell(row=1, column=1, value="Loan Purpose").fill = header_fill
ws_pivot.cell(row=1, column=1).font = header_font

for col_idx, col_name in enumerate(pivot_columns, 2):
    cell = ws_pivot.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font

# Write pivot data
for row_idx, (purpose, row_data) in enumerate(pivot_loan_purpose.iterrows(), 2):
    ws_pivot.cell(row=row_idx, column=1, value=purpose)
    for col_idx, value in enumerate(row_data, 2):
        ws_pivot.cell(row=row_idx, column=col_idx, value=value)

# Adjust column widths
ws_pivot.column_dimensions['A'].width = 20
for col_idx in range(2, len(pivot_columns) + 2):
    ws_pivot.column_dimensions[get_column_letter(col_idx)].width = 18

# Add employment status summary
employment_row = len(pivot_loan_purpose) + 3
ws_pivot.cell(row=employment_row, column=1, value="Employment Status").fill = header_fill
ws_pivot.cell(row=employment_row, column=1).font = header_font

employment_summary = df.groupby('employment_status').agg({
    'annual_income': 'mean',
    'loan_amount': 'mean',
    'credit_score': 'mean'
}).round(2)

emp_columns = list(employment_summary.columns)
for col_idx, col_name in enumerate(emp_columns, 2):
    cell = ws_pivot.cell(row=employment_row, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font

for row_idx, (emp_status, row_data) in enumerate(employment_summary.iterrows(), employment_row + 1):
    ws_pivot.cell(row=row_idx, column=1, value=emp_status)
    for col_idx, value in enumerate(row_data, 2):
        ws_pivot.cell(row=row_idx, column=col_idx, value=value)

# Add charts
print("\nCreating Charts...")

# Chart 1: Bar chart - Average Loan Amount by Loan Purpose
ws_chart1 = wb.create_sheet(title="Charts")
ws_chart1.sheet_view.tabSelected = False

# Prepare data for chart
chart_data_start_row = 1
chart_data = df.groupby('loan_purpose')['loan_amount'].mean().sort_values(ascending=False)

# Write chart data
ws_chart1.cell(row=1, column=1, value="Loan Purpose")
ws_chart1.cell(row=1, column=2, value="Avg Loan Amount")
for row_idx, (purpose, amount) in enumerate(chart_data.items(), 2):
    ws_chart1.cell(row=row_idx, column=1, value=purpose)
    ws_chart1.cell(row=row_idx, column=2, value=round(amount, 2))

# Create bar chart
bar_chart = BarChart()
bar_chart.title = "Average Loan Amount by Purpose"
bar_chart.style = 10
bar_chart.y_axis.title = "Average Loan Amount ($)"
bar_chart.x_axis.title = "Loan Purpose"

data = Reference(ws_chart1, min_col=2, min_row=1, max_row=len(chart_data) + 1)
cats = Reference(ws_chart1, min_col=1, min_row=2, max_row=len(chart_data) + 1)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(cats)
bar_chart.height = 10
bar_chart.width = 20

ws_chart1.add_chart(bar_chart, "D1")

# Chart 2: Histogram of Credit Scores
ws_chart1.cell(row=1, column=4, value="Credit Score")
for row_idx, score in enumerate(df['credit_score'].sort_values(), 2):
    ws_chart1.cell(row=row_idx, column=4, value=score)

# Create histogram using bar chart (manual bins)
# Bin the credit scores
credit_scores = df['credit_score'].sort_values()
bins = [500, 550, 600, 650, 700, 750, 800, 850, 900]
hist_data = pd.cut(credit_scores, bins=bins).value_counts().sort_index()

# Write histogram data
ws_chart1.cell(row=20, column=1, value="Credit Score Range")
ws_chart1.cell(row=20, column=2, value="Count")
for row_idx, (bin_range, count) in enumerate(hist_data.items(), 21):
    ws_chart1.cell(row=row_idx, column=1, value=str(bin_range))
    ws_chart1.cell(row=row_idx, column=2, value=count)

# Create bar chart for histogram
hist_chart = BarChart()
hist_chart.type = "col"
hist_chart.title = "Credit Score Distribution"
hist_chart.style = 11
hist_chart.y_axis.title = "Count"
hist_chart.x_axis.title = "Credit Score Range"

data_hist = Reference(ws_chart1, min_col=2, min_row=20, max_row=len(hist_data) + 20)
cats_hist = Reference(ws_chart1, min_col=1, min_row=21, max_row=len(hist_data) + 20)
hist_chart.add_data(data_hist, titles_from_data=True)
hist_chart.set_categories(cats_hist)
hist_chart.height = 10
hist_chart.width = 20

ws_chart1.add_chart(hist_chart, "D20")

# Chart 3: Loan count by Employment Status (Pie Chart)
ws_chart1.cell(row=1, column=6, value="Employment Status")
ws_chart1.cell(row=1, column=7, value="Count")
emp_counts = df['employment_status'].value_counts()

for row_idx, (status, count) in enumerate(emp_counts.items(), 2):
    ws_chart1.cell(row=row_idx, column=6, value=status)
    ws_chart1.cell(row=row_idx, column=7, value=count)

# Create pie chart
pie_chart = PieChart()
pie_chart.title = "Loans by Employment Status"
pie_chart.style = 10

data_pie = Reference(ws_chart1, min_col=7, min_row=1, max_row=len(emp_counts) + 1)
cats_pie = Reference(ws_chart1, min_col=6, min_row=2, max_row=len(emp_counts) + 1)
pie_chart.add_data(data_pie, titles_from_data=True)
pie_chart.set_categories(cats_pie)
pie_chart.height = 10
pie_chart.width = 15

ws_chart1.add_chart(pie_chart, "J1")

# Chart 4: Average Interest Rate by Loan Purpose
ws_chart1.cell(row=20, column=6, value="Loan Purpose")
ws_chart1.cell(row=20, column=7, value="Avg Interest Rate")
interest_data = df.groupby('loan_purpose')['interest_rate'].mean().sort_values(ascending=False)

for row_idx, (purpose, rate) in enumerate(interest_data.items(), 21):
    ws_chart1.cell(row=row_idx, column=6, value=purpose)
    ws_chart1.cell(row=row_idx, column=7, value=round(rate, 2))

# Create bar chart for interest rates
bar_chart2 = BarChart()
bar_chart2.title = "Average Interest Rate by Purpose"
bar_chart2.style = 12
bar_chart2.y_axis.title = "Interest Rate (%)"
bar_chart2.x_axis.title = "Loan Purpose"

data2 = Reference(ws_chart1, min_col=7, min_row=20, max_row=len(interest_data) + 20)
cats2 = Reference(ws_chart1, min_col=6, min_row=21, max_row=len(interest_data) + 20)
bar_chart2.add_data(data2, titles_from_data=True)
bar_chart2.set_categories(cats2)
bar_chart2.height = 10
bar_chart2.width = 18

ws_chart1.add_chart(bar_chart2, "J20")

# Save the workbook
output_path = f'{output_dir}/report.xlsx'
wb.save(output_path)
print(f"\n✓ Report saved to: {output_path}")

# Print summary
print("\n" + "="*60)
print("REPORT SUMMARY")
print("="*60)
print(f"Total records processed: {len(df)}")
print(f"\nSheets created:")
for sheet_name in sheet_configs.keys():
    print(f"  - {sheet_name}")
print(f"  - Summary (Pivot Table)")
print(f"  - Charts")
print(f"\nConditional formatting applied to outlier cells (IQR method)")
print(f"Charts added: 4 (Bar charts, Histogram, Pie chart)")
print("="*60)